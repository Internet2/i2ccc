"""
Weekly conversation export for ABE.

Runs Mondays at 8:00 AM Eastern (EventBridge Scheduler). Each run:

  1. reads the watermark - the newest message timestamp covered by the previous
     export - from SSM Parameter Store
  2. scans the conversation-history table and keeps everything newer than the
     watermark. With no watermark stored the whole history is exported, so the
     first email carries a complete (larger) workbook and later ones only the
     new week.
  3. builds an .xlsx that keeps every DynamoDB attribute as its own column, so
     the reader filters in Excel instead of asking for a different export
  4. uploads it and emails a presigned download link over SNS
  5. advances the watermark only after SNS accepts the message, so a failed
     send is picked up by the next run instead of silently losing a week

Two details worth knowing before changing this:

* Feedback is written onto the original message row (thumb_rating,
  feedback_text) without changing its timestamp, so a timestamp window alone
  would miss a thumbs-down left on an older answer. Every rated message
  therefore ships in its own sheet on every run.
* A DynamoDB FilterExpression is applied after the rows are read, so filtering
  the scan server-side would cost exactly the same as reading the table and
  splitting the rows here. One full scan feeds both sheets.
"""
import io
import json
import os
import re
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from urllib.parse import quote
from zoneinfo import ZoneInfo

import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

TABLE_NAME = os.environ.get("CONVERSATION_TABLE", "")
EXPORT_BUCKET = os.environ.get("EXPORT_BUCKET", "")
EXPORT_PREFIX = os.environ.get("EXPORT_PREFIX", "conversation-exports")
SNS_TOPIC_ARN = os.environ.get("SNS_TOPIC_ARN", "")
WATERMARK_PARAM = os.environ.get("WATERMARK_PARAM", "")
URL_EXPIRY_DAYS = int(os.environ.get("URL_EXPIRY_DAYS", "7"))
# 0 or unset means exports are kept indefinitely
RETENTION_DAYS = int(os.environ.get("EXPORT_RETENTION_DAYS", "0"))
REGION = os.environ.get("AWS_REGION", "us-east-1")
REPORT_TIMEZONE = os.environ.get("REPORT_TIMEZONE", "America/New_York")

dynamodb = boto3.resource("dynamodb")
s3 = boto3.client("s3")
sns = boto3.client("sns")
ssm = boto3.client("ssm")

RULE = "-" * 68
NEXT_RUN = "Mondays at 8:00 AM Eastern"
XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Excel refuses these control characters and truncates a cell past 32767 chars
_ILLEGAL_CELL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
MAX_CELL_CHARS = 32767

# Written on every run even when this week's rows never use them, so the column
# layout - and therefore any Excel filter or pivot built on it - stays put.
BASE_COLUMNS = (
    "session_id",
    "timestamp",
    "role",
    "content",
    "conversation_turn",
    "owner_sub",
    "thumb_rating",
    "feedback_text",
    "document_ids",
    "sources",
)

# Added by the export for readability; the raw attributes above are all kept
DERIVED_COLUMNS = (
    "datetime_eastern",
    "date_eastern",
    "datetime_utc",
    "source_count",
    "sources_titles",
)

COLUMN_WIDTHS = {
    "datetime_eastern": 20,
    "date_eastern": 13,
    "datetime_utc": 22,
    "source_count": 8,
    "sources_titles": 60,
    "session_id": 38,
    "timestamp": 16,
    "role": 11,
    "content": 90,
    "conversation_turn": 38,
    "owner_sub": 38,
    "thumb_rating": 14,
    "feedback_text": 24,
    "document_ids": 40,
    "sources": 60,
}

# Excel shows a datetime as a raw serial number without one of these
NUMBER_FORMATS = {
    "datetime_eastern": "yyyy-mm-dd hh:mm",
    "date_eastern": "yyyy-mm-dd",
}

CONVERSATIONS_SHEET = "conversations"
FEEDBACK_SHEET = "all_feedback"
RUN_INFO_SHEET = "run_info"


def _report_tz():
    try:
        return ZoneInfo(REPORT_TIMEZONE)
    except Exception as e:
        # tzdata is a dependency, but a missing zone must not lose the export
        print(f"Could not load timezone {REPORT_TIMEZONE}, using UTC: {e}")
        return timezone.utc


REPORT_TZ = _report_tz()


#####################################################################
# VALUE CONVERSION
#####################################################################
def _plain(value):
    """Make a DynamoDB-decoded value JSON-serializable."""
    if isinstance(value, Decimal):
        number = float(value)
        return int(number) if number.is_integer() else number
    if isinstance(value, dict):
        return {key: _plain(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_plain(item) for item in value]
    if isinstance(value, set):
        return sorted(_plain(item) for item in value)
    if isinstance(value, (bytes, bytearray)):
        return value.decode("utf-8", errors="replace")
    return value


def _cell(value):
    """Render one attribute as something Excel will accept."""
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, Decimal):
        return _plain(value)
    if isinstance(value, (int, float, datetime)):
        return value
    if isinstance(value, (list, tuple, dict, set)):
        value = json.dumps(_plain(value), ensure_ascii=False, default=str)
    text = _ILLEGAL_CELL_CHARS.sub("", str(value))
    if len(text) > MAX_CELL_CHARS:
        text = text[: MAX_CELL_CHARS - 15] + " ...[truncated]"
    return text


def _as_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _eastern(timestamp_ms):
    """Epoch milliseconds -> naive local datetime (Excel rejects tz-aware)."""
    if timestamp_ms is None:
        return None
    moment = datetime.fromtimestamp(timestamp_ms / 1000, tz=timezone.utc)
    return moment.astimezone(REPORT_TZ).replace(tzinfo=None)


def _utc_iso(timestamp_ms):
    if timestamp_ms is None:
        return None
    moment = datetime.fromtimestamp(timestamp_ms / 1000, tz=timezone.utc)
    return moment.strftime("%Y-%m-%d %H:%M:%S UTC")


def _sources_titles(item):
    sources = item.get("sources")
    if not isinstance(sources, list):
        return None
    titles = [
        str(source.get("title")).strip()
        for source in sources
        if isinstance(source, dict) and source.get("title")
    ]
    # A single answer often cites the same deck several times
    return " | ".join(dict.fromkeys(titles)) or None


def _derived(item):
    timestamp_ms = _as_int(item.get("timestamp"), default=None)
    local = _eastern(timestamp_ms)
    sources = item.get("sources")
    return {
        "datetime_eastern": local,
        "date_eastern": local.date() if local else None,
        "datetime_utc": _utc_iso(timestamp_ms),
        "source_count": len(sources) if isinstance(sources, list) else None,
        "sources_titles": _sources_titles(item),
    }


#####################################################################
# WORKBOOK
#####################################################################
def _columns_for(items):
    """Stable base layout first, then any attribute added to the table later."""
    seen = set()
    for item in items:
        seen.update(item.keys())
    extras = sorted(seen - set(BASE_COLUMNS))
    return list(DERIVED_COLUMNS) + list(BASE_COLUMNS) + extras


def _sorted_for_reading(items):
    """
    Group each conversation's messages together, conversations oldest first.

    Sorting on timestamp alone interleaves concurrent sessions, which makes a
    question and its answer hard to read side by side.
    """
    session_start = {}
    for item in items:
        session = str(item.get("session_id"))
        timestamp = _as_int(item.get("timestamp"))
        if session not in session_start or timestamp < session_start[session]:
            session_start[session] = timestamp
    return sorted(
        items,
        key=lambda item: (
            session_start.get(str(item.get("session_id")), 0),
            str(item.get("session_id")),
            _as_int(item.get("timestamp")),
        ),
    )


def _apply_column_styles(sheet, columns):
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(column, 20)
        number_format = NUMBER_FORMATS.get(column)
        if number_format:
            for cell in sheet[letter][1:]:
                cell.number_format = number_format


def _write_records_sheet(workbook, title, items, columns):
    sheet = workbook.create_sheet(title)
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for item in _sorted_for_reading(items):
        row = dict(_derived(item))
        for column in columns:
            if column not in DERIVED_COLUMNS:
                row[column] = item.get(column)
        sheet.append([_cell(row.get(column)) for column in columns])

    _apply_column_styles(sheet, columns)

    # Ready to filter the moment the file opens - the whole point of the export
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1:
        last = get_column_letter(len(columns))
        sheet.auto_filter.ref = f"A1:{last}{max(sheet.max_row, 1)}"
    return sheet


def _write_run_info_sheet(workbook, meta):
    sheet = workbook.create_sheet(RUN_INFO_SHEET)
    sheet.append(["field", "value"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for field, value in meta.items():
        sheet.append([field, _cell(value)])
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 70
    return sheet


def build_workbook(new_items, rated_items, meta):
    """Compose the workbook. Kept free of I/O so it can be tested directly."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    columns = _columns_for(list(new_items) + list(rated_items))
    _write_records_sheet(workbook, CONVERSATIONS_SHEET, new_items, columns)
    _write_records_sheet(workbook, FEEDBACK_SHEET, rated_items, columns)
    _write_run_info_sheet(workbook, meta)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


#####################################################################
# DATA
#####################################################################
def scan_all_messages():
    table = dynamodb.Table(TABLE_NAME)
    items = []
    kwargs = {}
    while True:
        page = table.scan(**kwargs)
        items += page.get("Items", [])
        last_key = page.get("LastEvaluatedKey")
        if not last_key:
            return items
        kwargs["ExclusiveStartKey"] = last_key


def load_watermark():
    if not WATERMARK_PARAM:
        return None
    try:
        value = ssm.get_parameter(Name=WATERMARK_PARAM)["Parameter"]["Value"]
        return int(value)
    except ssm.exceptions.ParameterNotFound:
        # First ever run: no watermark means "export everything"
        return None
    except (ClientError, TypeError, ValueError) as e:
        # A corrupt watermark would silently truncate the export, so fail loudly
        raise RuntimeError(f"Could not read watermark {WATERMARK_PARAM}: {e}") from e


def save_watermark(timestamp_ms):
    ssm.put_parameter(
        Name=WATERMARK_PARAM,
        Value=str(int(timestamp_ms)),
        Type="String",
        Overwrite=True,
        Description="Newest conversation timestamp covered by a weekly export",
    )


#####################################################################
# EMAIL
#####################################################################
def _figure(label, value, note=""):
    row = f"  {label.ljust(24)}{str(value).rjust(6)}"
    return f"{row}   {note}".rstrip()


def _window_phrase(watermark_ms, mode):
    if mode == "full":
        return "the complete history (first export)"
    start = _eastern(watermark_ms)
    if not start:
        return "the complete history"
    return f"messages after {start:%b} {start.day}, {start.year} {start:%H:%M} Eastern"


def _retention_lines():
    if RETENTION_DAYS:
        return [
            "  Save the file somewhere of your own if you want to keep it -",
            f"  copies in AWS are deleted after {RETENTION_DAYS} days.",
        ]
    return [
        "  Every past export is kept indefinitely, so the console links above",
        "  keep working however long ago the email was sent.",
    ]


def _console_object_url(key):
    """Console link to this one file - works after the presigned link expires."""
    return (
        f"https://{REGION}.console.aws.amazon.com/s3/object/{EXPORT_BUCKET}"
        f"?region={REGION}&bucketType=general&prefix={quote(key)}"
    )


def _console_prefix_url():
    """Console link to every export kept in the bucket."""
    return (
        f"https://{REGION}.console.aws.amazon.com/s3/buckets/{EXPORT_BUCKET}"
        f"?region={REGION}&bucketType=general&prefix={quote(EXPORT_PREFIX)}/"
    )


def build_email(stats, meta, url, filename):
    """Compose the (subject, body) pair. Kept separate from I/O for testing."""
    generated = stats["generated_eastern"]
    new_rows = stats["new_rows"]
    questions = stats["questions"]
    answers = stats["answers"]
    sessions = stats["sessions"]
    expires = stats["expires_eastern"]

    if new_rows:
        conversation_word = "conversation" if sessions == 1 else "conversations"
        headline = (
            f"{new_rows} new messages across {sessions} {conversation_word} "
            "- the Excel file is ready to download."
        )
    else:
        headline = (
            "No new questions were asked this week. The file still has every "
            "rating ever left, in case you want to review those."
        )

    lines = [
        "ABE CONVERSATION EXPORT",
        f"{generated:%A, %B} {generated.day}, {generated.year}",
        "",
        headline,
        "",
        "DOWNLOAD",
        f"  {url}",
        "",
        f"  THIS LINK EXPIRES IN {URL_EXPIRY_DAYS} DAYS, on "
        f"{expires:%A, %B} {expires.day}, {expires.year}.",
        f"  Covers {_window_phrase(stats['watermark_ms'], stats['mode'])}.",
        "",
        "AFTER THE LINK EXPIRES",
        "  The file itself is not going anywhere. Sign in to the AWS console",
        "  first, then open it directly:",
        "",
        f"  {_console_object_url(filename)}",
        "",
        "  Every past export is listed here:",
        "",
        f"  {_console_prefix_url()}",
        "",
        "  Both links need you to be signed in to AWS - open them in the same",
        "  browser where you are signed in, then use the Download button.",
        "",
        "THE BIG PICTURE",
        _figure("Questions asked", questions),
        _figure("Answers returned", answers),
        _figure("Conversations", sessions),
        _figure("Thumbs up (this batch)", stats["thumbs_up_new"]),
        _figure("Thumbs down (this batch)", stats["thumbs_down_new"]),
        _figure("Rated messages, all time", stats["rated_total"], "(sheet 2)"),
        "",
        "WHAT'S IN THE FILE",
        f"  Sheet 1  {CONVERSATIONS_SHEET.ljust(15)}"
        "every message in this export, one per row",
        f"  Sheet 2  {FEEDBACK_SHEET.ljust(15)}"
        "every thumbs up/down ever left, with its reason",
        f"  Sheet 3  {RUN_INFO_SHEET.ljust(15)}"
        "what this export covered, for the record",
        "",
        "HOW TO USE IT",
        "  Every field stored per message is its own column and filters are",
        "  already switched on, so you should not need a different export:",
        "",
        "    - filter role = user to read just the questions people asked",
        "    - filter thumb_rating = thumbs_down to see what fell short, with",
        "      the reason in feedback_text",
        "    - sources_titles lists the documents ABE cited in each answer",
        "    - a question and its answer share a conversation_turn value and",
        "      sit next to each other",
        "",
        "  Sheet 2 exists because a rating can be left days after the answer.",
        "  Those rows would fall outside this week's window, so every rated",
        "  message is repeated there in full.",
        "",
        *_retention_lines(),
        "",
        f"  The next export is sent {NEXT_RUN}.",
        "",
        RULE,
        "TECHNICAL DETAILS",
        "",
        f"File: s3://{EXPORT_BUCKET}/{filename}",
        f"Table: {TABLE_NAME}",
        f"Mode: {stats['mode']}",
        f"Watermark before this run: {stats['watermark_ms'] or 'unset (full export)'}",
        f"Watermark after this run: {stats['new_watermark_ms'] or 'unchanged'}",
        f"Rows scanned: {stats['scanned_rows']}",
        f"Rows written to sheet 1: {new_rows}",
        f"Rows written to sheet 2: {stats['rated_total']}",
        f"Presigned link requested for: {URL_EXPIRY_DAYS} days",
        "",
        "The presigned link is signed with this function's temporary",
        "credentials, so it can stop working before the stated expiry if those",
        "credentials rotate. The console links above are the fallback and need",
        "the reader to have console access to the export bucket.",
    ]
    for field, value in meta.items():
        if field.startswith("export_"):
            lines.append(f"{field}: {value}")

    subject = (
        f"ABE conversations: {new_rows} new messages"
        if new_rows
        else "ABE conversations: no new activity"
    )
    subject = f"{subject} - {generated:%b} {generated.day}"
    # SNS subjects must be ASCII and under 100 chars
    subject = subject.encode("ascii", errors="ignore").decode().strip()[:100]
    return subject, "\n".join(lines)


#####################################################################
# S3
#####################################################################
def _presign(key, filename):
    return s3.generate_presigned_url(
        "get_object",
        Params={
            "Bucket": EXPORT_BUCKET,
            "Key": key,
            # Without this the browser saves the file under its full key path
            "ResponseContentDisposition": (
                f'attachment; filename="{os.path.basename(filename)}"'
            ),
            "ResponseContentType": XLSX_CONTENT_TYPE,
        },
        ExpiresIn=URL_EXPIRY_DAYS * 86400,
    )


#####################################################################
# HANDLER
#####################################################################
def handler(event, context):
    event = event or {}
    full = bool(event.get("full"))
    watermark_ms = None if full else load_watermark()
    if event.get("since_ms") is not None:
        watermark_ms = int(event["since_ms"])
    mode = "full" if watermark_ms is None else "incremental"

    all_items = scan_all_messages()
    new_items = [
        item
        for item in all_items
        if watermark_ms is None or _as_int(item.get("timestamp")) > watermark_ms
    ]
    rated_items = [item for item in all_items if item.get("thumb_rating")]

    generated = datetime.now(timezone.utc)
    generated_local = generated.astimezone(REPORT_TZ).replace(tzinfo=None)
    timestamps = [_as_int(item.get("timestamp")) for item in new_items]
    new_watermark_ms = max(timestamps) if timestamps else watermark_ms

    stats = {
        "mode": mode,
        "generated_eastern": generated_local,
        "expires_eastern": generated_local + timedelta(days=URL_EXPIRY_DAYS),
        "watermark_ms": watermark_ms,
        "new_watermark_ms": new_watermark_ms if timestamps else None,
        "scanned_rows": len(all_items),
        "new_rows": len(new_items),
        "questions": sum(1 for item in new_items if item.get("role") == "user"),
        "answers": sum(1 for item in new_items if item.get("role") == "assistant"),
        "sessions": len({str(item.get("session_id")) for item in new_items}),
        "thumbs_up_new": sum(
            1 for item in new_items if item.get("thumb_rating") == "thumbs_up"
        ),
        "thumbs_down_new": sum(
            1 for item in new_items if item.get("thumb_rating") == "thumbs_down"
        ),
        "rated_total": len(rated_items),
    }

    meta = {
        "export_generated_eastern": f"{generated_local:%Y-%m-%d %H:%M}",
        "export_generated_utc": generated.strftime("%Y-%m-%d %H:%M:%S UTC"),
        "export_mode": mode,
        "export_covers": _window_phrase(watermark_ms, mode),
        "export_rows_sheet1": stats["new_rows"],
        "export_rows_sheet2": stats["rated_total"],
        "export_questions": stats["questions"],
        "export_answers": stats["answers"],
        "export_conversations": stats["sessions"],
        "export_table": TABLE_NAME,
        "export_watermark_before": watermark_ms or "unset (full export)",
        "export_watermark_after": new_watermark_ms or "unchanged",
    }

    workbook = build_workbook(new_items, rated_items, meta)
    stamp = generated.strftime("%Y%m%d-%H%M")
    key = f"{EXPORT_PREFIX}/abe-conversations-{stamp}.xlsx"
    s3.put_object(
        Bucket=EXPORT_BUCKET,
        Key=key,
        Body=workbook,
        ContentType=XLSX_CONTENT_TYPE,
    )
    print(f"Wrote {len(workbook)} bytes to s3://{EXPORT_BUCKET}/{key}")

    url = _presign(key, key)
    subject, body = build_email(stats, meta, url, key)
    sns.publish(TopicArn=SNS_TOPIC_ARN, Subject=subject, Message=body)

    # Only now is the week safely delivered; a send failure above leaves the
    # watermark alone so the next run re-exports this window.
    if timestamps and event.get("advance_watermark", True):
        save_watermark(new_watermark_ms)
        print(f"Watermark advanced to {new_watermark_ms}")

    return {
        "key": key,
        "bytes": len(workbook),
        "new_rows": stats["new_rows"],
        "rated_total": stats["rated_total"],
        "mode": mode,
        "watermark": new_watermark_ms,
    }
